"""
块探测器 — 报价汇总 sheet 自动块/子块识别.

输入: openpyxl Worksheet (data_only=False 推荐, data_only=True 也可)
输出: list of dicts:
    {
        "title": str,         # 块标题 (有 fallback 策略)
        "ds": int,            # 数据起始行 (1-indexed, 源行号)
        "de": int,            # 数据结束行 (1-indexed, 源行号)
        "hr": int,            # HEADER 行 (R2/R3/R20 等) ← 用于 smart_col_map
        "subblocks": [        # 子块列表 (副表头行切分)
            {"ds": int, "de": int},  # 源行号
            ...
        ],
        "first_a": str | None,        # 块内第一个 A 列非空值
        "series_names": [str, ...],   # 块内所有 SERIES 名 (去重, 按出现顺序)
    }

行识别规则:
- TITLE:  A 是字符串 + 含 "报价 20" 等强关键词 OR (无数据时)是 series 类字符串
- HEADER: A=="类别" or "SERIES" 且 C/D == "订单明细" or "原型机"
- DATA:   C 或 D 匹配 ^Z[0-9A-Z]{10,}$ (物料号)
- SUBHEADER (副表头): A/C/D/G 全空 + H/I/J/K 至少 2 列有文本 → 切分子块, 写入跳过
- SUBTOTAL (数量小计): A/C/D 空 + G 是数字 + J 空 → 完全跳过 (不计入 ds/de)
- TOTAL: A/C/D 空 + Col 18-24 含 "总" 字 或 =SUM( 公式 → 跳过 (用户决策)
- FRAGMENT: A 空但 C/D/G 是物料号/数字 → 属于当前块

切分子块的时机:
1. 副表头行 (上面定义)
2. SERIES 名在块内已经存在 → 该行作为新 V 系列 (不切 W 子块)

W 子块 (V 是 SERIES 级, W 是 子块级) 仅在副表头行切分。

设计要点:
- 输出 block["hr"] 字段 (HEADER 行), 供 smart_col_map_from_header 识别每块表头
- HR 字段支持 R3+ (如贸易商/埃及询盘 R1 空, R2 标题, R3 表头)
- 块间不共享 col_map: 每块独立从 hr 字段识别
- 详见 references/MIGRATION_RULES.md §7

注意: TOTAL 检测需要 data_only=False (看到 =SUM( 字符串); data_only=True 时
TOTAL 行的 =SUM( 会变成数字 (0), is_total_row_at() 仍能检测到 '总' 字。
"""
import re

PATTERN_Z = re.compile(r'^Z[0-9A-Z]{10,}$')
TITLE_KW_STRICT = [
    "报价 20", "客户订单", "客户报价", "首次报价", "询盘", "样机询盘",
    "冷年", "sets", "SETS", "返单", "翻单", "现场", "定版待回签",
]


def is_subheader_row(ws, r):
    """A/C/D/G 空 + H/I/J/K 至少 2 列有文本 → 副表头行 (子块分割 + 输出时跳过)"""
    a = ws.cell(r, 1).value
    c = ws.cell(r, 3).value
    d = ws.cell(r, 4).value
    g = ws.cell(r, 7).value
    if a is not None or c is not None or d is not None or g is not None:
        return False
    text_count = 0
    for col in range(8, 12):  # H, I, J, K
        v = ws.cell(r, col).value
        if v is not None and isinstance(v, str):
            text_count += 1
    return text_count >= 2


def is_quantity_subtotal_row(ws, r):
    """A/C/D 空 + G 有数字 + J 空 → 数量小计行 (探测 + 写入都跳过)"""
    a = ws.cell(r, 1).value
    c = ws.cell(r, 3).value
    d = ws.cell(r, 4).value
    g = ws.cell(r, 7).value
    j10 = ws.cell(r, 10).value  # 部分 sheet 的 J 列
    j11 = ws.cell(r, 11).value  # IBC/白鲸的 J 列
    if a is not None or c is not None or d is not None:
        return False
    if g is None or not isinstance(g, (int, float)):
        return False
    if j10 is not None or j11 is not None:
        return False
    return True


def is_total_row_at(ws, r, max_col=25):
    """
    总计行检测 (用户决定: 跳过不写).
    特征:
    - 1-17 列全空
    - 18-24 列任一含 '总' 字 或 =SUM( 公式

    注意: data_only=False 时 =SUM( 是字符串; data_only=True 时 =SUM( 是数字 (0).
    两种 mode 都能检测到, 因为 '总' 字面 + =SUM 至少一个会触发.
    """
    if r < 1 or r > ws.max_row:
        return False
    if any(ws.cell(r, c).value is not None for c in range(1, 18)):
        return False
    for c in range(18, max_col + 1):
        v = ws.cell(r, c).value
        if v == '总':
            return True
        if isinstance(v, str) and v.startswith('=SUM('):
            return True
    return False


def detect_blocks(ws):
    """
    自动探测块的标题行、表头行、数据范围、子块分割点。
    返回: list of dicts (见模块顶部说明)

    使用注意:
    - 推荐 data_only=False (确保 =SUM( 字符串被识别为 TOTAL)
    - 但 data_only=True 也能工作 (TOTAL 通过 '总' 字面识别)
    - 用户决策: TOTAL 行被 is_total_row_at 识别后 continue 跳过, 不纳入任何块
    """
    blocks = []
    cur = {
        "title": None, "ds": None, "de": None, "hr": None,
        "subblocks": [], "first_a": None, "series_names": [],
    }
    last_title = None
    cur_sub_ds = None

    def flush_block():
        nonlocal cur, cur_sub_ds
        flush_sub()
        if cur["ds"] is not None:
            # 标题 fallback 链 (优先级从高到低)
            if not cur["title"]:
                if cur.get("series_names"):
                    short_names = []
                    for sn in cur["series_names"][:2]:
                        first_line = sn.split("\n")[0].strip()
                        if first_line and first_line not in short_names:
                            short_names.append(first_line)
                    if len(cur["series_names"]) > 2:
                        cur["title"] = " / ".join(short_names) + " ..."
                    else:
                        cur["title"] = " / ".join(short_names)
                elif cur.get("first_a"):
                    cur["title"] = cur["first_a"].split("\n")[0].strip()
                else:
                    cur["title"] = last_title or "(未命名块)"
            blocks.append(dict(cur))
        cur = {
            "title": None, "ds": None, "de": None, "hr": None,
            "subblocks": [], "first_a": None, "series_names": [],
        }
        cur_sub_ds = None

    def flush_sub():
        nonlocal cur, cur_sub_ds
        if cur_sub_ds is not None and cur["de"] is not None and cur_sub_ds <= cur["de"]:
            cur["subblocks"].append({"ds": cur_sub_ds, "de": cur["de"]})
        cur_sub_ds = None

    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        c = ws.cell(r, 3).value
        d = ws.cell(r, 4).value
        g = ws.cell(r, 7).value

        # 总计行检测 (必须在 all_blank 之前! 否则 R44 等会先被当空行 flush)
        if is_total_row_at(ws, r):
            if cur["ds"] is not None:
                cur["de"] = r  # 把总计行并入当前块 (用户决定不写, 但块范围需要包含它)
            continue

        all_blank = all(ws.cell(r, c2).value is None for c2 in range(1, 9))
        if all_blank:
            flush_block()
            continue

        if a in ("类别", "SERIES") and (c in ("订单明细", "原型机") or d in ("订单明细", "原型机")):
            cur["ds"] = r + 1
            cur["de"] = r + 1
            cur["hr"] = r
            cur_sub_ds = r + 1
            continue

        if is_subheader_row(ws, r):
            flush_sub()
            cur_sub_ds = r + 1
            continue

        if is_quantity_subtotal_row(ws, r):
            continue  # 完全跳过

        is_z = False
        if c and PATTERN_Z.match(str(c).strip()):
            is_z = True
        elif d and PATTERN_Z.match(str(d).strip()):
            is_z = True

        if is_z:
            if cur["ds"] is None:
                continue
            cur["de"] = r
            if cur["first_a"] is None and a:
                cur["first_a"] = str(a).strip()
            if a and isinstance(a, str) and a.strip() not in cur["series_names"]:
                cur["series_names"].append(a.strip())
            continue

        if a and isinstance(a, str):
            is_strong = any(kw in a for kw in TITLE_KW_STRICT)
            is_series = (
                " " in a
                and not any(ch in a for ch in "()（）")
                and "TEL" not in a
                and not a.startswith("Z")
            )
            if is_strong or (cur["ds"] is None and is_series):
                flush_block()
                cur["title"] = a.strip()
                last_title = cur["title"]
                continue

        if (c is not None or d is not None or (g is not None and isinstance(g, (int, float)))) and cur["ds"] is not None:
            cur["de"] = r
            continue

    flush_block()
    return blocks


if __name__ == "__main__":
    import openpyxl
    import sys

    if len(sys.argv) < 2:
        print("用法: python detect_blocks.py <xlsx文件> [sheet名]")
        sys.exit(2)

    src = sys.argv[1]
    wb = openpyxl.load_workbook(src, data_only=True)
    sheets = [sys.argv[2]] if len(sys.argv) > 2 else wb.sheetnames

    for sn in sheets:
        if sn not in wb.sheetnames:
            print(f"  跳过 (不存在): {sn}")
            continue
        ws = wb[sn]
        blocks = detect_blocks(ws)
        print(f"\n【{sn}】{len(blocks)} 块:")
        for i, b in enumerate(blocks):
            sb_desc = ", ".join(f"R{sb['ds']}-{sb['de']}" for sb in b["subblocks"]) or "(整块)"
            print(f"  块{i+1}: title='{b['title'][:35]}' | R{b['ds']}-R{b['de']} | hr=R{b.get('hr')} | 子块: {sb_desc}")
    wb.close()
