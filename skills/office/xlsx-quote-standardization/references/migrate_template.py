# ============================================================
# 报价汇总迁移模板
# 用法: 复制本文件到工作目录 → 修改 CONFIG 区 → 运行
# 依赖: openpyxl（SKILL.md 已声明）
# ============================================================
import sys
import os
_SKILL_DIR = os.path.dirname(os.path.abspath(__file__))
if _SKILL_DIR not in sys.path:
    sys.path.insert(0, _SKILL_DIR)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from detect_blocks import detect_blocks, is_subheader_row, is_quantity_subtotal_row

# ============================================================
# CONFIG — 用户只改这里
# ============================================================
SRC = r"<源文件路径>"
OUT = r"<输出文件路径>"

# 跳过的 sheet (哲学层/空 sheet)
SKIP_SHEETS = {"333", "FRESH转口"}

# 源 sheet -> 输出 sheet 名映射
SHEET_OUTPUT_NAMES = {
    "埃及机型报价模板": "01_埃及机型", "军方项目报价": "02_军方项目", "IBC": "03_IBC",
    "白鲸报价": "04_白鲸", "INDIGO": "05_INDIGO", "FROSTIL": "06_FROSTIL",
    "NICE AIR": "07_NICE_AIR", "TRUMAN": "08_TRUMAN", "Fresh拖多": "09_Fresh拖多",
    "TRK": "10_TRK", "FRESH本土": "11_FRESH本土", "ELB ": "12_ELB",
    "ETS": "13_ETS", "LEGIM": "14_LEGIM", "iclima": "15_iclima",
    "贸易商": "16_贸易商", "MXP": "17_MXP", "埃及询盘": "18_埃及询盘",
    "询盘": "19_询盘",
}

# ============================================================
# STYLES
# ============================================================
FONT_TITLE    = Font(name='Arial', size=12, bold=True, color='000000')
FONT_HEADER   = Font(name='Arial', size=10, bold=True, color='FFFFFF')
FONT_BODY     = Font(name='Arial', size=10)
FONT_FORMULA  = Font(name='Arial', size=10, italic=True, color='008000')
FONT_INPUT    = Font(name='Arial', size=10, color='0000FF')
FONT_AGGREGATE= Font(name='Arial', size=10, bold=True, color='C00000')
FILL_TITLE    = PatternFill('solid', fgColor='D9E1F2')
FILL_HEADER   = PatternFill('solid', fgColor='4472C4')
FILL_INPUT    = PatternFill('solid', fgColor='FFF2CC')
FILL_FORMULA  = PatternFill('solid', fgColor='F2F2F2')
FILL_AGGREGATE= PatternFill('solid', fgColor='E2EFDA')
FILL_PENDING  = PatternFill('solid', fgColor='FCE4D6')
ALIGN_LEFT    = Alignment(horizontal='left', vertical='center', wrapText=True)
ALIGN_RIGHT   = Alignment(horizontal='right', vertical='center')
ALIGN_CENTER  = Alignment(horizontal='center', vertical='center', wrapText=True)
BORDER_THIN   = Border(left=Side(style='thin',color='BFBFBF'),right=Side(style='thin',color='BFBFBF'),
                        top=Side(style='thin',color='BFBFBF'),bottom=Side(style='thin',color='BFBFBF'))
FMT_USD = r'\$#,##0.00_);[Red]\(\$#,##0.00\)'
FMT_USD_MINUS = r'\$#,##0.00;\-\$#,##0.00'
FMT_PCT = '0.00%'
COL_FMT = {1:'General',2:'General',3:'General',4:'General',5:'General',6:'General',7:'General',
    8:FMT_USD,9:FMT_USD,10:FMT_USD,11:FMT_USD,12:FMT_USD,13:FMT_USD,14:FMT_USD,15:FMT_USD,
    16:FMT_USD,17:FMT_USD,18:FMT_USD,19:FMT_USD,20:FMT_USD_MINUS,21:FMT_PCT,22:FMT_PCT,23:FMT_PCT,24:FMT_USD}
HEADERS_24 = ['类别','产品类别','订单明细','工厂型号','配置描述','压缩机','数量',
              '历史报价','客户目标价','报价',
              '财务费用','OA信保','返点','其他费用','净价','原型机成本','铜管成本(3m)','结算价',
              '净收入','毛利',
              '单机型损益率','系列盈亏','总盈亏','铜管规格']
COL_WIDTHS = {1:38,2:28,3:19,4:38,5:38,6:22,7:9,8:15,9:15,10:15,11:14,12:14,13:14,14:14,15:15,
              16:16,17:15,18:16,19:18,20:18,21:12,22:12,23:12,24:22}

# ============================================================
# SMART COL_MAP (核心)
# ============================================================
def smart_col_map_from_header(ws, header_row):
    """从源表头自动识别 col_map. 目标布局固定 24 列."""
    cells = {c: ws.cell(header_row, c).value for c in range(1, 26)}

    def find_col(predicate, exclude_words=()):
        for c in range(8, 25):
            v = cells.get(c)
            if isinstance(v, str) and predicate(v) and not any(w in v for w in exclude_words):
                return c
        return None

    # 报价列 (3 级 fallback)
    quote_col = None
    for c in range(8, 25):
        if cells.get(c) == "报价":
            quote_col = c; break
    if not quote_col:
        for c in range(8, 25):
            if cells.get(c) == "PRICE":
                quote_col = c; break
    # 第 3 级: '价格' (iclima 等 sheet 表头用 '价格' 而非 '报价')
    if not quote_col:
        for c in range(8, 25):
            if cells.get(c) == "价格":
                quote_col = c; break
    if not quote_col:
        exclude = ("历史报价", "客户目标价", "客户最近目标价", "目标价", "原报价",
                   "上次报价", "涨跌幅", "涨幅", "4/13报价", "最近一次报价",
                   "上一次报价", "上一次报价（92000/7.0）", "科特迪瓦成交价报价", "TRK成交价（含管）",
                   "向ELB终端报价")
        for c in range(8, 25):
            v = cells.get(c)
            if isinstance(v, str) and "报价" in v and v not in exclude:
                quote_col = c; break

    # 净价/原型机/铜管/结算价 (按字面匹配)
    n_price_col = find_col(lambda v: v == "净价")
    proto_col   = find_col(lambda v: v.startswith("原型机成本"))
    copper_col  = find_col(lambda v: v.startswith("铜管成本"))
    settle_col  = find_col(lambda v: v == "结算价")

    # K/L/M/N (按字面精确匹配)
    k_col = find_col(lambda v: v == "财务费用")
    l_col = find_col(lambda v: v == "OA信保")
    m_col = find_col(lambda v: v == "返点")
    n_col = find_col(lambda v: v == "其他费用")
    if not n_col:
        n_col = find_col(lambda v: v == "其他成本")

    # S (净收入/总金额) 和 T (毛利)
    s_col = find_col(lambda v: v in ("净收入", "总金额"))
    t_col = find_col(lambda v: v == "毛利")

    # 铜管规格/连接管线 (Col 24/25)
    extra_col = None
    for c in (24, 25):
        v = cells.get(c)
        if v and ("铜管" in str(v) or "连接" in str(v)):
            extra_col = c; break

    col_map = {1:1, 2:2, 3:3, 4:4, 5:5, 6:6, 7:7}
    if quote_col:  col_map[quote_col] = 10
    if k_col:      col_map[k_col] = 11
    if l_col:      col_map[l_col] = 12
    if m_col:      col_map[m_col] = 13
    if n_col:      col_map[n_col] = 14
    if proto_col:  col_map[proto_col] = 16
    if copper_col: col_map[copper_col] = 17
    if settle_col: col_map[settle_col] = 18
    if extra_col:  col_map[extra_col] = 24

    sheet_meta = {"s_col": s_col, "t_col": t_col, "n_price_col": n_price_col}
    return col_map, sheet_meta

# ============================================================
# HELPERS
# ============================================================
def safe_num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace('$','').replace(',','').replace(' ','').replace('\xa0','')
    try: return float(s)
    except: return None

def sc(cell, font=None, fill=None, alignment=None, nf=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if nf: cell.number_format = nf
    cell.border = BORDER_THIN

def write_data_rows(ws_out, start_row, src_ws, src_start, src_end, col_map, sheet_meta=None):
    """
    写数据行。
    - 副表头行 (A/C/D/G 空 + H/I/J/K 含文本) 跳过不写
    - 数量小计行已由探测过滤
    - 总计行已由探测跳过 (用户决定)
    """
    ds = start_row
    out_idx = 0
    actual_de = None
    for src_r in range(src_start, src_end + 1):
        if is_subheader_row(src_ws, src_r):
            continue
        out_r = ds + out_idx
        for tc in range(1, 25):
            sc_src = None
            for s, t in col_map.items():
                if t == tc:
                    sc_src = s
                    break
            cell = ws_out.cell(row=out_r, column=tc)
            raw = src_ws.cell(src_r, sc_src).value if sc_src else None
            if tc in (1, 2, 3, 4, 5, 6):
                cell.value = str(raw).strip() if raw else None
                sc(cell, font=FONT_BODY, alignment=ALIGN_LEFT if tc == 1 else ALIGN_RIGHT)
            elif tc == 7:
                num = safe_num(raw)
                cell.value = num if (num and num > 0) else None
                sc(cell, font=FONT_BODY, alignment=ALIGN_RIGHT)
            elif tc in (8, 9):
                cell.value = safe_num(raw)
                sc(cell, font=FONT_BODY, alignment=ALIGN_RIGHT, nf=COL_FMT[tc])
            elif tc == 10:
                cell.value = safe_num(raw)
                sc(cell, font=FONT_INPUT, fill=FILL_INPUT, alignment=ALIGN_RIGHT, nf=COL_FMT[tc])
            elif tc in (11, 12, 13, 14):
                num = safe_num(raw)
                cell.value = num if num is not None else 0
                sc(cell, font=FONT_BODY, alignment=ALIGN_RIGHT, nf=COL_FMT[tc])
            elif tc == 15:
                # O 净价 = J - K - L - M - N (铁律)
                cell.value = f'=IFERROR(J{out_r}-K{out_r}-L{out_r}-M{out_r}-N{out_r},0)'
                sc(cell, font=FONT_FORMULA, fill=FILL_FORMULA, alignment=ALIGN_RIGHT, nf=COL_FMT[tc])
            elif tc == 16:
                num = safe_num(raw)
                cell.value = num if (num and num > 0) else None
                sc(cell, font=FONT_BODY, alignment=ALIGN_RIGHT, nf=COL_FMT[tc])
                if cell.value is None:
                    cell.fill = FILL_PENDING
            elif tc == 17:
                num = safe_num(raw)
                cell.value = num if num is not None else 0
                sc(cell, font=FONT_BODY, alignment=ALIGN_RIGHT, nf=COL_FMT[tc])
            elif tc == 18:
                cell.value = f'=IFERROR(P{out_r}+Q{out_r},0)'
                sc(cell, font=FONT_FORMULA, fill=FILL_FORMULA, alignment=ALIGN_RIGHT, nf=COL_FMT[tc])
            elif tc == 19:
                # S 净收入 = O*G (基于净价, 不是 J*G)
                cell.value = f'=IFERROR(O{out_r}*G{out_r},0)'
                sc(cell, font=FONT_FORMULA, fill=FILL_FORMULA, alignment=ALIGN_RIGHT, nf=COL_FMT[tc])
            elif tc == 20:
                # T 毛利 = (O-R)*G (基于净价, 不是 (J-R)*G)
                cell.value = f'=IFERROR((O{out_r}-R{out_r})*G{out_r},0)'
                sc(cell, font=FONT_FORMULA, fill=FILL_FORMULA, alignment=ALIGN_RIGHT, nf=COL_FMT[tc])
            elif tc == 21:
                cell.value = f'=IFERROR(IF(S{out_r}=0,0,T{out_r}/S{out_r}),0)'
                sc(cell, font=FONT_FORMULA, fill=FILL_FORMULA, alignment=ALIGN_RIGHT, nf=COL_FMT[tc])
            elif tc in (22, 23):
                sc(cell, font=FONT_AGGREGATE, fill=FILL_AGGREGATE, alignment=ALIGN_CENTER, nf=FMT_PCT)
            elif tc == 24:
                num = safe_num(raw)
                cell.value = num if num is not None else (str(raw).strip() if raw else 0)
                sc(cell, font=FONT_BODY, alignment=ALIGN_RIGHT, nf=COL_FMT[tc])
        a_val = ws_out.cell(out_r, 1).value
        if a_val and '\n' in str(a_val):
            lines = str(a_val).count('\n') + 1
            ws_out.row_dimensions[out_r].height = max(15, lines * 14)
        actual_de = out_r
        out_idx += 1
    return ds, actual_de if actual_de is not None else ds - 1

def write_title_header(ws_out, row, title):
    cell = ws_out.cell(row=row, column=1, value=title)
    sc(cell, font=FONT_TITLE, fill=FILL_TITLE, alignment=ALIGN_CENTER)
    ws_out.merge_cells(start_row=row, start_column=1, end_row=row, end_column=24)
    hr = row + 1
    for c in range(1, 25):
        cell = ws_out.cell(row=hr, column=c, value=HEADERS_24[c-1])
        sc(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER)
    return hr

def apply_vw_with_subblocks(ws_out, ds, de, subblocks):
    """
    Apply V/W formulas. V 按系列分组合并, W 按子块分组合并.
    subblocks: [{"ds": r1, "de": r2}, ...] 输出行号
    """
    # 找 A 列系列边界
    series = []
    cn, cs = None, None
    for r in range(ds, de + 1):
        a = ws_out.cell(r, 1).value
        if a:
            if cn: series.append((cn, cs, r - 1))
            cn, cs = a, r
    if cn: series.append((cn, cs, de))

    # V 列按系列合并
    for name, s, e in series:
        f = f'=IFERROR(SUM(T{s}:T{e})/SUM(S{s}:S{e}),0)'
        c = ws_out.cell(row=s, column=22, value=f)
        sc(c, font=FONT_AGGREGATE, fill=FILL_AGGREGATE, alignment=ALIGN_CENTER, nf=FMT_PCT)
        if e > s:
            ws_out.merge_cells(start_row=s, start_column=22, end_row=e, end_column=22)

    # W 列按子块合并
    if not subblocks:
        subblocks = [{"ds": ds, "de": de}]
    for sb in subblocks:
        s, e = sb["ds"], sb["de"]
        f = f'=IFERROR(SUM(T{s}:T{e})/SUM(S{s}:S{e}),0)'
        c = ws_out.cell(row=s, column=23, value=f)
        sc(c, font=FONT_AGGREGATE, fill=FILL_AGGREGATE, alignment=ALIGN_CENTER, nf=FMT_PCT)
        if e > s:
            ws_out.merge_cells(start_row=s, start_column=23, end_row=e, end_column=23)

    # A 列按系列合并
    for name, s, e in series:
        if e > s:
            ws_out.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)

def fix_column_widths(ws):
    for col_num, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width

# ============================================================
# MAIN
# ============================================================
print("Loading source...")
wb_src_f = load_workbook(SRC, data_only=False)  # 探测 (看公式字面值)
wb_src_v = load_workbook(SRC, data_only=True)   # 读缓存值

wb_out = Workbook()
wb_out.remove(wb_out.active)

ALL_SRC_SHEETS = [sn for sn in SHEET_OUTPUT_NAMES.keys() if sn not in SKIP_SHEETS]
summary = []

for src_name in ALL_SRC_SHEETS:
    out_name = SHEET_OUTPUT_NAMES[src_name]
    ws_src_f = wb_src_f[src_name]
    ws_src_v = wb_src_v[src_name]
    ws_out = wb_out.create_sheet(out_name)

    blocks = detect_blocks(ws_src_f)
    print(f"\n  {out_name:18s} <- {src_name:14s} | {len(blocks)} 块")

    cur_row = 1
    block_summaries = []
    for i, blk in enumerate(blocks):
        title = blk["title"]
        ds, de = blk["ds"], blk["de"]
        src_subblocks = blk.get("subblocks", [])

        # 智能生成 col_map (每块独立, 从该块 header_row 自动识别)
        if "hr" in blk and blk["hr"]:
            col_map, sheet_meta = smart_col_map_from_header(ws_src_f, blk["hr"])
        else:
            col_map, sheet_meta = smart_col_map_from_header(ws_src_f, 2)

        hr = write_title_header(ws_out, cur_row, title)
        # 注意: write_data_rows 必须用 data_only=True (ws_src_v) 读缓存值,
        # 因为源 P/Q 列常含公式 (=4.66*3, =102%*153.17 等),
        # data_only=False 读到公式字符串 safe_num 无法解析 → 输出全 0
        ds_out, de_out = write_data_rows(ws_out, hr + 1, ws_src_v, ds, de, col_map, sheet_meta=sheet_meta)

        # 重新算 src_to_out (副表头行已跳过)
        # 用 ws_src_v 保持与 write_data_rows 一致
        src_to_out = {}
        out_idx = 0
        for src_r in range(ds, de + 1):
            if is_subheader_row(ws_src_v, src_r):
                continue
            src_to_out[src_r] = ds_out + out_idx
            out_idx += 1

        # 转换子块到输出行号
        out_subblocks = []
        for sb in src_subblocks:
            if sb["ds"] in src_to_out and sb["de"] in src_to_out:
                out_subblocks.append({"ds": src_to_out[sb["ds"]], "de": src_to_out[sb["de"]]})
        apply_vw_with_subblocks(ws_out, ds_out, de_out, out_subblocks)

        sb_desc = ", ".join(f"R{sb['ds']}-{sb['de']}" for sb in out_subblocks) if out_subblocks else "整块"
        block_summaries.append(f"R{ds}-R{de} (输出R{ds_out}-R{de_out}, 子:{sb_desc})")
        cur_row = de_out + 2

    fix_column_widths(ws_out)
    print(f"    块: {'; '.join(block_summaries)}")
    summary.append({"src": src_name, "out": out_name, "blocks": len(blocks)})

print(f"\nSaving {OUT}...")
wb_out.save(OUT)
wb_out.close()
wb_src_f.close()
wb_src_v.close()

print(f"\n完成! 共迁移 {len(ALL_SRC_SHEETS)} 个 sheet\n摘要:")
for s in summary:
    print(f"  {s['out']:18s} <- {s['src']:14s} | {s['blocks']} 块")
