"""
语义化对账脚本。

用 smart_col_map_from_header 反查每块 col_map, 不需要 SRC_COLS 硬编码.
每行按"语义列"含义比对: J(报价)/G(数量)/P(原型机)/Q(铜管)/S(净收入)/T(毛利).

核价场景:
  "在进行核价时，是否应该多关注语义而非行列的对应关系呢？"

用法:
    python verify_semantic.py <源.xlsx> <输出.xlsx>

设计:
- 用 detect_blocks 找到每块的 ds/de/hr
- smart_col_map_from_header 从每块 hr 字段识别 col_map
- 反查 target col -> src col (t_to_s = {t: s for s, t in col_map.items()})
- 按 C 列物料号顺序匹配源行 -> 输出行
- 排除副表头行 (is_subheader_row)
"""
import openpyxl
import re
import sys
import os

# 允许从 references/ 目录引用 detect_blocks.py
_SKILL_DIR = os.path.dirname(os.path.abspath(__file__))
_REFS_DIR = os.path.join(os.path.dirname(_SKILL_DIR), "references")
for p in (_REFS_DIR, _SKILL_DIR):
    if p not in sys.path:
        sys.path.insert(0, p)

from detect_blocks import detect_blocks, is_subheader_row


PATTERN_Z = re.compile(r'^Z[0-9A-Z]{10,}$')


def smart_col_map_from_header(ws, header_row):
    """从源表头自动识别 col_map. 核心函数. 复制自 migrate_template.py."""
    cells = {c: ws.cell(header_row, c).value for c in range(1, 26)}

    def find_col(predicate, exclude_words=()):
        for c in range(8, 25):
            v = cells.get(c)
            if isinstance(v, str) and predicate(v) and not any(w in v for w in exclude_words):
                return c
        return None

    quote_col = None
    for c in range(8, 25):
        if cells.get(c) == "报价":
            quote_col = c; break
    if not quote_col:
        for c in range(8, 25):
            if cells.get(c) == "PRICE":
                quote_col = c; break
    # 第 3 级: '价格' (iclima 等 sheet 表头用 '价格' 而非 '报价')
    if not quote_col:
        for c in range(8, 25):
            if cells.get(c) == "价格":
                quote_col = c; break
    if not quote_col:
        exclude = ("历史报价", "客户目标价", "客户最近目标价", "目标价", "原报价",
                   "上次报价", "涨跌幅", "涨幅", "4/13报价", "最近一次报价",
                   "上一次报价", "上一次报价（92000/7.0）", "科特迪瓦成交价报价", "TRK成交价（含管）",
                   "向ELB终端报价")
        for c in range(8, 25):
            v = cells.get(c)
            if isinstance(v, str) and "报价" in v and v not in exclude:
                quote_col = c; break

    n_price_col = find_col(lambda v: v == "净价")
    proto_col   = find_col(lambda v: v.startswith("原型机成本"))
    copper_col  = find_col(lambda v: v.startswith("铜管成本"))
    settle_col  = find_col(lambda v: v == "结算价")

    k_col = find_col(lambda v: v == "财务费用")
    l_col = find_col(lambda v: v == "OA信保")
    m_col = find_col(lambda v: v == "返点")
    n_col = find_col(lambda v: v == "其他费用")
    if not n_col:
        n_col = find_col(lambda v: v == "其他成本")

    s_col = find_col(lambda v: v in ("净收入", "总金额"))
    t_col = find_col(lambda v: v == "毛利")

    extra_col = None
    for c in (24, 25):
        v = cells.get(c)
        if v and ("铜管" in str(v) or "连接" in str(v)):
            extra_col = c; break

    col_map = {1:1, 2:2, 3:3, 4:4, 5:5, 6:6, 7:7}
    if quote_col:  col_map[quote_col] = 10
    if k_col:      col_map[k_col] = 11
    if l_col:      col_map[l_col] = 12
    if m_col:      col_map[m_col] = 13
    if n_col:      col_map[n_col] = 14
    if proto_col:  col_map[proto_col] = 16
    if copper_col: col_map[copper_col] = 17
    if settle_col: col_map[settle_col] = 18
    if extra_col:  col_map[extra_col] = 24

    sheet_meta = {"s_col": s_col, "t_col": t_col, "n_price_col": n_price_col}
    return col_map, sheet_meta


def safe_num(v):
    if v is None:
        return None
    if isinstance(v, str):
        if v.startswith("#"):
            return None
        s = v.strip().replace("$", "").replace(",", "").replace(" ", "").replace("\xa0", "")
        try:
            return float(s)
        except ValueError:
            return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


# 输出永远是标准 24 列布局
OUT_COLS = {"J": 10, "G": 7, "P": 16, "Q": 17, "S": 19, "T": 20}


def main(src_path, out_path, src_to_out_map=None):
    """
    src_to_out_map: {源 sheet 名: 输出 sheet 名} 映射
    默认: 01_xxx, 02_xxx, ... 编号格式
    """
    if src_to_out_map is None:
        src_to_out_map = {
            "埃及机型报价模板": "01_埃及机型", "军方项目报价": "02_军方项目", "IBC": "03_IBC",
            "白鲸报价": "04_白鲸", "INDIGO": "05_INDIGO", "FROSTIL": "06_FROSTIL",
            "NICE AIR": "07_NICE_AIR", "TRUMAN": "08_TRUMAN", "Fresh拖多": "09_Fresh拖多",
            "TRK": "10_TRK", "FRESH本土": "11_FRESH本土", "ELB ": "12_ELB",
            "ETS": "13_ETS", "LEGIM": "14_LEGIM", "iclima": "15_iclima",
            "贸易商": "16_贸易商", "MXP": "17_MXP", "埃及询盘": "18_埃及询盘",
            "询盘": "19_询盘",
        }

    wb_s = openpyxl.load_workbook(src_path, data_only=True)
    wb_o = openpyxl.load_workbook(out_path, data_only=True)

    sheet_results = []
    total_correct = 0
    total_wrong = 0
    total_skipped = 0
    wrong_samples = []

    for src_name, out_name in src_to_out_map.items():
        if src_name not in wb_s.sheetnames:
            continue
        if out_name not in wb_o.sheetnames:
            continue

        ws_s = wb_s[src_name]
        ws_o = wb_o[out_name]

        # 探测所有块
        blocks = detect_blocks(ws_s)

        # 输出 sheet 的所有数据行 (C 列含 Z 物料号)
        out_data_rows = []
        for r in range(1, ws_o.max_row + 1):
            c_val = ws_o.cell(r, 3).value
            if c_val and PATTERN_Z.match(str(c_val).strip()):
                out_data_rows.append(r)

        out_idx = 0
        sheet_ok = 0
        sheet_wrong = 0
        for blk in blocks:
            ds, de = blk["ds"], blk["de"]
            hr = blk.get("hr")

            if not hr:
                continue

            # 智能生成 col_map (每块独立)
            col_map, _ = smart_col_map_from_header(ws_s, hr)
            # 反向: 目标 col -> 源 col
            t_to_s = {t: s for s, t in col_map.items()}

            for src_r in range(ds, de + 1):
                if is_subheader_row(ws_s, src_r):
                    continue

                c_val = ws_s.cell(src_r, 3).value
                if c_val is None or not PATTERN_Z.match(str(c_val).strip()):
                    continue

                if out_idx >= len(out_data_rows):
                    total_skipped += 1
                    continue
                out_r = out_data_rows[out_idx]
                out_idx += 1

                # 按语义列比对
                row_diffs = []
                for semantic, out_col in OUT_COLS.items():
                    src_col = t_to_s.get(out_col)
                    if src_col is None:
                        # 源没有这一列, 跳过 (例如净收入 S 不在源)
                        continue
                    src_v = ws_s.cell(src_r, src_col).value
                    out_v = ws_o.cell(out_r, out_col).value
                    sn, on = safe_num(src_v), safe_num(out_v)
                    if (sn or 0) != (on or 0):
                        row_diffs.append(f"{semantic}=src{src_col}:{sn}≠out{out_col}:{on}")

                if not row_diffs:
                    sheet_ok += 1
                    total_correct += 1
                else:
                    sheet_wrong += 1
                    total_wrong += 1
                    if len(wrong_samples) < 20:
                        wrong_samples.append(f"  ✗ {out_name} R{src_r}({c_val}) → R{out_r}: {row_diffs}")

        sheet_results.append((src_name, sheet_ok, sheet_wrong, out_name))

    # 打印结果
    print("="*100)
    print(f"语义化逐行对账 (智能 col_map)")
    print("="*100)
    print(f"\n{'源 Sheet':25s} | {'数据行':>6s} | {'正确':>6s} | {'差异':>6s} | {'正确率':>8s}")
    print("-"*70)
    for sn, ok, wrong, out_name in sheet_results:
        total = ok + wrong
        rate = ok * 100 / total if total else 0
        print(f"  {sn:25s} | {total:>6d} | {ok:>6d} | {wrong:>6d} | {rate:>7.1f}%")

    grand_total = total_correct + total_wrong
    rate = total_correct * 100 / grand_total if grand_total else 0
    print(f"\n  {'总计':25s} | {grand_total:>6d} | {total_correct:>6d} | {total_wrong:>6d} | {rate:>7.1f}%")
    print(f"  跳过: {total_skipped}")

    if wrong_samples:
        print(f"\n差异样本 (前 20):")
        for s in wrong_samples:
            print(s)

    wb_s.close()
    wb_o.close()

    return 0 if total_wrong == 0 else 1


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("用法: python verify_semantic.py <源.xlsx> <输出.xlsx>")
        sys.exit(2)
    sys.exit(main(sys.argv[1], sys.argv[2]))
