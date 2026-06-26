"""
逐列对账脚本 — 对比源 vs 输出在每一列的有效值总和。

为什么需要这个脚本:
  verify_totals.py 只对 J 列做"块总和"对比。它能抓漏块/全 0 偏移，
  但抓不到"col_map 漏配某一列"的情况。

  案例: INDIGO 块 3/4 的 col_map 漏写 14:11 (财务费用), 15:12 (OA信保),
         16:13 (返点) — 源数据 K=1.25 L=0.31 M=0 P=132 都有值,
         但输出 K=0 L=0 M=0 → J-K-L-M-N = J, 净价公式算错。
         J 列总和验证通过 (J 没变), 但公式列 O 全部偏 1.5% 左右。

使用:
    python verify_columns.py <源.xlsx> <输出.xlsx>

输出: 每行一个 (sheet, column) 的 src_sum vs out_sum, 任何差异 > 1.0 标 ✗。
"""
import openpyxl
import sys
from collections import defaultdict


# 每个 sheet 的列映射配置: {src_sheet: {out_sheet: {src_col: out_col, ...}}}
# 这里只列已知 6 个 sheet 的关键列; 新增 sheet 时补上
COLUMN_MAPS = {
    # 源 sheet -> {输出 sheet -> {(src_col, out_col): label, ...}}
    "埃及机型报价模板": {
        "01_埃及机型": [
            (10, 10, "J 报价"),
            (14, 11, "K 财务"), (15, 12, "L OA"), (16, 13, "M 返点"), (17, 16, "P 原型机"),
            (18, 17, "Q 铜管"), (19, 19, "S 净收入(应=0)"), (24, 24, "X 铜管规格"),
        ],
        # 块 2 用 IBC-like 映射 (源 11=PRICE)
        "_块2": [
            (11, 10, "J 报价"), (12, 11, "K 财务"), (13, 12, "L OA"), (14, 13, "M 返点"),
            (17, 16, "P 原型机"),
        ],
    },
    "军方项目报价": {"02_军方项目": [(10, 10, "J"), (14, 11, "K"), (15, 12, "L"), (17, 16, "P")]},
    "IBC":         {"03_IBC":     [(11, 10, "J"), (12, 11, "K"), (13, 12, "L"), (14, 13, "M"),
                                    (17, 16, "P"), (18, 17, "Q"), (25, 24, "X")]},
    "白鲸报价":     {"04_白鲸":    [(11, 10, "J"), (12, 11, "K"), (13, 12, "L"), (14, 13, "M"),
                                    (17, 16, "P"), (18, 17, "Q"), (25, 24, "X")]},
    "INDIGO":      {"05_INDIGO":  [(12, 10, "J"), (14, 11, "K"), (15, 12, "L"), (16, 13, "M"),
                                    (17, 16, "P"), (18, 17, "Q"), (25, 24, "X")]},
    "FROSTIL":     {"06_FROSTIL": [(12, 10, "J"), (14, 11, "K"), (15, 12, "L"), (16, 13, "M"),
                                    (17, 16, "P"), (18, 17, "Q"), (25, 24, "X")]},
}

# 数据块范围 (ds, de) — 与 verify_totals.py 保持一致
SHEET_BLOCKS = {
    "埃及机型报价模板": [("01_埃及机型", [(3, 52, 10), (86, 102, 11)])],
    "军方项目报价":     [("02_军方项目", [(3, 5, 10)])],
    "IBC":             [("03_IBC", [(3, 15, 11), (21, 43, 11)])],
    "白鲸报价":         [("04_白鲸", [(3, 15, 11), (22, 35, 11)])],
    "INDIGO":          [("05_INDIGO", [(3, 8, 12), (14, 14, 12), (19, 29, 12), (34, 44, 12)])],
    "FROSTIL":         [("06_FROSTIL", [(3, 8, 12), (14, 22, 12), (28, 31, 12), (36, 42, 12), (49, 55, 12)])],
}


def src_sum_col(ws, col, blocks):
    """某 sheet 多个块在 col 列上的有效数字总和"""
    s = 0
    for ds, de in blocks:
        for r in range(ds, de + 1):
            v = ws.cell(r, col).value
            if isinstance(v, (int, float)):
                s += v
    return s


def out_data_rows(ws):
    """输出 sheet 的所有数据行 (排除标题行 + 表头行)"""
    rows = []
    title_rows = []
    for r in range(1, ws.max_row + 1):
        for mr in ws.merged_cells.ranges:
            if mr.min_row == r and mr.min_col == 1 and mr.max_col == 24:
                title_rows.append(r); break
    for i, tr in enumerate(title_rows):
        first = tr + 2
        last = title_rows[i + 1] - 2 if i + 1 < len(title_rows) else ws.max_row
        rows.extend(range(first, last + 1))
    return rows


def out_sum_col(ws, col, data_rows):
    """输出 sheet 在 col 列上的有效数字总和"""
    s = 0
    for r in data_rows:
        v = ws.cell(r, col).value
        if isinstance(v, (int, float)):
            s += v
    return s


def main(src_path, out_path):
    wb_s = openpyxl.load_workbook(src_path, data_only=True)
    wb_o = openpyxl.load_workbook(out_path, data_only=True)

    failures = 0
    print(f"{'sheet':16s} {'col':18s} {'src sum':>12s} {'out sum':>12s}  status")
    print("-" * 75)

    for src_name, sheet_cfg in COLUMN_MAPS.items():
        if src_name not in wb_s.sheetnames:
            continue
        ws_s = wb_s[src_name]

        for out_name, col_list in sheet_cfg.items():
            if out_name.startswith("_"):
                continue
            if out_name not in wb_o.sheetnames:
                continue
            ws_o = wb_o[out_name]
            data_rows = out_data_rows(ws_o)

            blocks_list = SHEET_BLOCKS.get(src_name, [])
            blocks = []
            for on, bl in blocks_list:
                if on == out_name:
                    blocks = [b[:2] for b in bl]  # 去掉 j_col
                    break

            for src_col, out_col, label in col_list:
                src_s = src_sum_col(ws_s, src_col, blocks)
                out_s = out_sum_col(ws_o, out_col, data_rows)
                diff = abs(src_s - out_s)
                status = "✓" if diff < 1.0 else "✗"
                if diff >= 1.0:
                    failures += 1
                print(f"{out_name:16s} {label:18s} {src_s:12.2f} {out_s:12.2f}  {status} (diff={diff:.2f})")

    wb_s.close(); wb_o.close()
    print()
    if failures:
        print(f"✗ {failures} 列数据不一致 — 通常是 col_map 漏配, 检查脚本里的 SHEET_SPECS[col_maps]")
        return 1
    print(f"✓ 所有列数据一致")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("用法: python verify_columns.py <源.xlsx> <输出.xlsx>")
        sys.exit(2)
    sys.exit(main(sys.argv[1], sys.argv[2]))
