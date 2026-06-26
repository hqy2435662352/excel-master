"""
数据对账脚本 — 对比源 vs 输出的 J 列（报价）总和，验证迁移数据完整性。

使用：
    python verify_totals.py <源.xlsx> <输出.xlsx>

配置：J_SOURCES 字典列出每个 (源 sheet, 输出 sheet, 源 J 列号, 数据块范围)。
若某个 sheet 的 J 列总和差异 > 1.0，标 ✗ 列出明细。
"""
import openpyxl
import sys
from pathlib import Path

J_SOURCES = [
    # (src_name, out_name, j_col_in_src, [(ds, de), ...])
    ("埃及机型报价模板", "01_埃及机型", None, [(3, 52, 10), (86, 102, 11)]),  # 块 1 用 Col10, 块 2 用 Col11
    ("IBC", "03_IBC", 11, [(3, 15), (21, 43)]),
    ("白鲸报价", "04_白鲸", 11, [(3, 15), (22, 35)]),
    ("INDIGO", "05_INDIGO", 12, [(3, 8), (14, 14), (19, 29), (34, 44)]),
    ("FROSTIL", "06_FROSTIL", 12, [(3, 8), (14, 22), (28, 31), (36, 42), (49, 55)]),
    ("NICE AIR", "07_NICE_AIR", 11, [(3, 12), (26, 31)]),
    ("TRUMAN", "08_TRUMAN", 11, [(3, 9), (12, 25), (33, 39)]),
    ("Fresh拖多", "09_Fresh拖多", 10, [(3, 6), (13, 21), (29, 37)]),
    ("TRK", "10_TRK", 10, [(3, 6), (12, 15), (21, 24)]),
    ("FRESH本土", "11_FRESH本土", 10, [(3, 5), (12, 18), (25, 27)]),
    ("ELB ", "12_ELB", 9, [(3, 5), (12, 15), (20, 25), (34, 35), (42, 43)]),
    ("ETS", "13_ETS", 9, [(3, 9)]),
    ("LEGIM", "14_LEGIM", 9, [(3, 6), (11, 14), (20, 24), (30, 32), (40, 41)]),
    ("iclima", "15_iclima", 9, [(3, 6), (10, 13), (18, 21), (27, 33), (40, 42), (49, 49), (54, 55), (61, 62), (66, 68)]),
    ("贸易商", "16_贸易商", 9, [(4, 10), (15, 18), (32, 35)]),
    ("MXP", "17_MXP", 9, [(3, 9)]),
    ("埃及询盘", "18_埃及询盘", 10, [(4, 9), (16, 31)]),
    ("询盘", "19_询盘", 8, [(3, 5)]),
    ("军方项目报价", "02_军方项目", 10, [(3, 5)]),
]


def src_sum(ws, j_col, blocks):
    s, c = 0, 0
    if j_col is not None:
        for ds, de in blocks:
            for r in range(ds, de + 1):
                v = ws.cell(r, j_col).value
                if isinstance(v, (int, float)):
                    s += v; c += 1
    else:
        for ds, de, col in blocks:
            for r in range(ds, de + 1):
                v = ws.cell(r, col).value
                if isinstance(v, (int, float)):
                    s += v; c += 1
    return s, c


def out_sum(ws):
    """输出 sheet 的 J 列：每个标题行（merged A:X）+2 到下一标题行 -2"""
    s, c = 0, 0
    title_rows = []
    for r in range(1, ws.max_row + 1):
        for mr in ws.merged_cells.ranges:
            if mr.min_row == r and mr.min_col == 1 and mr.max_col == 24:
                title_rows.append(r)
                break
    for i, tr in enumerate(title_rows):
        first = tr + 2
        last = title_rows[i + 1] - 2 if i + 1 < len(title_rows) else ws.max_row
        for r in range(first, last + 1):
            v = ws.cell(r, 10).value
            if isinstance(v, (int, float)):
                s += v; c += 1
    return s, c


def main(src_path, out_path):
    wb_src = openpyxl.load_workbook(src_path, data_only=True)
    wb_out = openpyxl.load_workbook(out_path, data_only=True)

    print(f"源: {src_path}")
    print(f"输出: {out_path}\n")
    print(f"{'sheet':18s} {'src sum':>14s} {'src cnt':>8s} {'out sum':>14s} {'out cnt':>8s}  status")
    print("-" * 80)

    failures = 0
    for src_name, out_name, j_col, blocks in J_SOURCES:
        s, c = src_sum(wb_src[src_name], j_col, blocks)
        o_s, o_c = out_sum(wb_out[out_name])
        diff = abs(s - o_s)
        status = "✓" if diff < 1.0 else "✗"
        if diff >= 1.0:
            failures += 1
        print(f"{out_name:18s} {s:14.2f} {c:8d} {o_s:14.2f} {o_c:8d}  {status} (diff={diff:.2f})")

    wb_src.close(); wb_out.close()
    print()
    if failures:
        print(f"✗ {failures} 个 sheet 数据不一致")
        return 1
    print(f"✓ 所有 sheet J 列总和一致")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("用法: python verify_totals.py <源.xlsx> <输出.xlsx>")
        sys.exit(2)
    sys.exit(main(sys.argv[1], sys.argv[2]))
